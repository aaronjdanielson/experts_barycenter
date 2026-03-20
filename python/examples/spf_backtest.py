"""
SPF Historical Backtest: Arithmetic Mean vs. W1 Barycenter
===========================================================

Evaluates aggregate PRCCPI probability forecasts against realized
Q4/Q4 core CPI outcomes using Ranked Probability Score (RPS) and
log score, for every quarter from 2007:Q1 through 2024:Q4.

Both SA and NSA realized CPI series are downloaded to verify the
target definition.  Results are split by cross-sectional dispersion
of individual forecast means (high vs. low).

Interpreter: /Users/aarondanielson/miniforge3/envs/supercluster/bin/python3.11
"""

import sys
import numpy as np
import pandas as pd
import openpyxl
from pathlib import Path
from collections import defaultdict
import matplotlib.pyplot as plt
import warnings
warnings.filterwarnings("ignore")

# ── Paths ────────────────────────────────────────────────────────────────────

ROOT = Path(__file__).parent.parent.parent
DATA_PATH = ROOT / "data" / "spf" / "SPFmicrodata.xlsx"
OUT_DIR   = ROOT / "output" / "spf_backtest"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ── FRED API key ─────────────────────────────────────────────────────────────
# Set FRED_API_KEY here or as an environment variable.
import os
FRED_API_KEY = os.environ.get("FRED_API_KEY", "")

# ── Bin definitions (ascending inflation order) ───────────────────────────────
# PRCCPI1-10 (SPF descending) are reversed to ascending in load_quarter().
#   Bin 0: will decline (< 0%)
#   Bin 1: 0.0 – 0.4%
#   Bin 2: 0.5 – 0.9%
#   ...
#   Bin 9: >= 4.0%

BIN_LABELS = [
    "Decline", "0.0–0.4", "0.5–0.9", "1.0–1.4", "1.5–1.9",
    "2.0–2.4", "2.5–2.9", "3.0–3.4", "3.5–3.9", "≥4.0",
]
K = 10

# Upper bounds of each bin (inf for last)
BIN_UPPERS = [-0.0, 0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, float("inf")]

def realized_bin(value: float) -> int:
    """Map a Q4/Q4 percent change to a bin index (0-based, ascending)."""
    if value < 0:
        return 0
    for i, upper in enumerate(BIN_UPPERS[1:], start=1):
        if value < upper:
            return i
    return K - 1


# ── Load PRCCPI data ─────────────────────────────────────────────────────────

def load_quarter(ws, header, year: int, qtr: int,
                 col_slice: slice, tol: float = 2.0) -> np.ndarray | None:
    """
    Return (n, K) array of valid forecaster histograms for (year, qtr).
    col_slice: slice(0,10) for current-year, slice(10,20) for next-year.
    Rows whose probability sum deviates more than `tol` from 100 are dropped.
    Remaining rows are renormalized to sum exactly to 1.
    Returns None if fewer than 5 valid forecasters remain.
    """
    prob_cols = [c for c in header if c and str(c).startswith("PRCCPI")]
    sel_cols  = prob_cols[col_slice]

    valid = []
    for row in ws.iter_rows(values_only=True):
        if row[0] != year or row[1] != qtr:
            continue
        vals = []
        for c in sel_cols:
            v = row[header.index(c)]
            if v in (None, "#N/A", ""):
                vals.append(0.0)
            else:
                try:
                    vals.append(float(v))
                except (TypeError, ValueError):
                    vals.append(0.0)
        s = sum(vals)
        if abs(s - 100.0) <= tol and s > 0:
            arr = np.array(vals) / s
            valid.append(arr[::-1])   # reverse to ascending order

    if len(valid) < 5:
        return None
    return np.array(valid)


def load_all_quarters(horizon: str = "current") -> dict:
    """
    Returns dict mapping (year, qtr) -> (n, K) ndarray.
    horizon: "current" uses PRCCPI1-10; "next" uses PRCCPI11-20.
    """
    col_slice = slice(0, 10) if horizon == "current" else slice(10, 20)
    wb = openpyxl.load_workbook(DATA_PATH, read_only=True)
    ws = wb["PRCCPI"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    quarters = sorted(set(
        (int(r[0]), int(r[1])) for r in rows[1:] if r[0] is not None
    ))

    data = {}
    for (yr, qt) in quarters:
        if yr < 2007:
            continue
        mat = load_quarter(ws, header, yr, qt, col_slice)
        if mat is not None:
            data[(yr, qt)] = mat
    wb.close()
    return data


# ── CDF-median barycenter ────────────────────────────────────────────────────

def cdf_median(pmfs: np.ndarray) -> np.ndarray:
    """Component-wise CDF median → PMF."""
    cdfs   = np.cumsum(pmfs, axis=1)
    med    = np.median(cdfs, axis=0)
    pmf    = np.diff(np.concatenate([[0.0], med]))
    pmf    = np.maximum(pmf, 0.0)
    pmf   /= pmf.sum()
    return pmf


def log_pool(pmfs: np.ndarray, eps: float = 1e-10) -> np.ndarray:
    """
    Logarithmic opinion pool (geometric mean, renormalized).
    LP(k) ∝ prod_j p_j(k)^{1/N} = exp(mean_j log p_j(k)).
    Clips to eps before log to avoid -inf.
    """
    log_mean = np.mean(np.log(np.maximum(pmfs, eps)), axis=0)
    unnorm   = np.exp(log_mean - log_mean.max())   # subtract max for stability
    return unnorm / unnorm.sum()


def trimmed_mean(pmfs: np.ndarray, trim: float = 0.10) -> np.ndarray:
    """
    Symmetric trimmed arithmetic mean.
    For each bin, removes the bottom and top `trim` fraction of values
    before averaging.  Uses scipy.stats.trim_mean under the hood.
    """
    from scipy.stats import trim_mean as _trim_mean
    result = np.array([_trim_mean(pmfs[:, k], trim) for k in range(pmfs.shape[1])])
    result = np.maximum(result, 0.0)
    s = result.sum()
    return result / s if s > 0 else np.ones(pmfs.shape[1]) / pmfs.shape[1]


# ── Scoring ──────────────────────────────────────────────────────────────────

def rps(pmf: np.ndarray, bin_idx: int) -> float:
    """Ranked Probability Score (lower = better)."""
    cdf_f = np.cumsum(pmf)
    cdf_r = np.zeros(K)
    cdf_r[bin_idx:] = 1.0
    return float(np.sum((cdf_f[:-1] - cdf_r[:-1]) ** 2))


def log_score(pmf: np.ndarray, bin_idx: int, eps: float = 1e-8) -> float:
    """Log score (higher = better)."""
    return float(np.log(max(pmf[bin_idx], eps)))


# ── Download realized CPI from FRED ─────────────────────────────────────────

def fetch_q4q4_cpi(series_id: str) -> pd.Series:
    """
    Download monthly CPI level from FRED and compute Q4/Q4
    (December-over-December) percent changes.
    Returns a Series indexed by year (int).
    """
    if not FRED_API_KEY:
        raise RuntimeError(
            "FRED API key not set.  Export FRED_API_KEY=<your_key> or set it "
            "at the top of this script."
        )
    from fredapi import Fred
    fred = Fred(api_key=FRED_API_KEY)
    monthly = fred.get_series(series_id, observation_start="2006-01-01",
                               observation_end="2025-12-31")
    monthly = monthly.dropna()
    # December values only
    dec = monthly[monthly.index.month == 12]
    pct = dec.pct_change() * 100
    pct = pct.dropna()
    # Index by year of the ending December
    return pd.Series(pct.values, index=pct.index.year, name=series_id)


# ── Forecast mean (for dispersion measure) ───────────────────────────────────

BIN_MIDPOINTS = np.array([-0.5, 0.2, 0.75, 1.25, 1.75, 2.25, 2.75, 3.25, 3.75, 4.5])

def forecast_mean(pmf: np.ndarray) -> float:
    return float(pmf @ BIN_MIDPOINTS)

def cross_sectional_variance(pmfs: np.ndarray) -> float:
    means = pmfs @ BIN_MIDPOINTS
    return float(np.var(means, ddof=1))


# ── Main backtest ─────────────────────────────────────────────────────────────

def run_backtest(horizon: str = "current"):
    print(f"\n{'='*60}")
    print(f"SPF Backtest  |  Horizon: {horizon}-year")
    print(f"{'='*60}")

    # Load SPF data
    panel = load_all_quarters(horizon)
    print(f"Quarters loaded: {len(panel)}")

    # Download realized CPI (both SA and NSA)
    print("\nDownloading realized CPI from FRED...")
    try:
        r_sa  = fetch_q4q4_cpi("CPILFESL")   # Seasonally adjusted
        r_nsa = fetch_q4q4_cpi("CPILFENS")   # Not seasonally adjusted
        print(f"  SA  (CPILFESL): {r_sa.index[0]}–{r_sa.index[-1]}")
        print(f"  NSA (CPILFENS): {r_nsa.index[0]}–{r_nsa.index[-1]}")
    except RuntimeError as e:
        print(f"\nFRED download failed: {e}")
        print("Using hardcoded NSA Q4/Q4 core CPI values (2007–2024).")
        # Source: BLS CPI for All Urban Consumers, All Items Less Food &
        # Energy, December-over-December, not seasonally adjusted.
        hardcoded = {
            2007: 2.4, 2008: 1.8, 2009: 1.7, 2010: 0.8,
            2011: 2.2, 2012: 1.9, 2013: 1.7, 2014: 1.6,
            2015: 2.1, 2016: 2.2, 2017: 1.8, 2018: 2.2,
            2019: 2.3, 2020: 1.6, 2021: 5.0, 2022: 5.7,
            2023: 3.9, 2024: 3.2,
        }
        r_sa  = pd.Series(hardcoded, name="CPILFESL_hardcoded")
        r_nsa = r_sa.copy()

    results = []

    for (yr, qt), pmfs in sorted(panel.items()):
        # For current-year horizon: forecast made in (yr, qt), realized = yr
        # For next-year horizon: forecast made in (yr, qt), realized = yr+1
        target_year = yr if horizon == "current" else yr + 1

        for series_label, r_series in [("SA", r_sa), ("NSA", r_nsa)]:
            if target_year not in r_series.index:
                continue
            realized_val = r_series[target_year]
            bin_idx      = realized_bin(realized_val)

            am  = pmfs.mean(axis=0)
            bc  = cdf_median(pmfs)
            lp  = log_pool(pmfs)
            tm  = trimmed_mean(pmfs, trim=0.10)
            csv = cross_sectional_variance(pmfs)
            n   = len(pmfs)

            row = {
                "year": yr, "qtr": qt, "target_year": target_year,
                "series": series_label, "n_forecasters": n,
                "realized": realized_val, "bin_idx": bin_idx,
                "am_rps":   rps(am, bin_idx),
                "bc_rps":   rps(bc, bin_idx),
                "lp_rps":   rps(lp, bin_idx),
                "tm_rps":   rps(tm, bin_idx),
                "am_ls":    log_score(am, bin_idx),
                "bc_ls":    log_score(bc, bin_idx),
                "lp_ls":    log_score(lp, bin_idx),
                "tm_ls":    log_score(tm, bin_idx),
                "am_mass":  am[bin_idx],
                "bc_mass":  bc[bin_idx],
                "lp_mass":  lp[bin_idx],
                "tm_mass":  tm[bin_idx],
                "dispersion": csv,
            }
            # Save full PMF for reliability diagram
            for k in range(K):
                row[f"am_p{k}"] = am[k]
                row[f"bc_p{k}"] = bc[k]
                row[f"lp_p{k}"] = lp[k]
                row[f"tm_p{k}"] = tm[k]
            results.append(row)

    df = pd.DataFrame(results)
    df.to_csv(OUT_DIR / f"backtest_{horizon}.csv", index=False)
    print(f"\nSaved raw results → output/spf_backtest/backtest_{horizon}.csv")

    for series_label in ["SA", "NSA"]:
        sub = df[df["series"] == series_label].copy()
        if sub.empty:
            continue
        _print_summary(sub, horizon, series_label)
        _plot_scores(sub, horizon, series_label)
        _plot_dispersion_split(sub, horizon, series_label)
        if horizon == "current":
            _plot_reliability(sub, series_label)

    return df


def _dm_test(loss_a, loss_b):
    """Diebold-Mariano test, H1: loss_a < loss_b (one-sided)."""
    from scipy import stats as _stats
    d = np.array(loss_a) - np.array(loss_b)
    n = len(d)
    d_bar = d.mean()
    var_d = np.var(d, ddof=0)
    se = np.sqrt(max(var_d, 1e-30) / n)
    dm = d_bar / se
    return float(dm), float(_stats.norm.cdf(dm))


def _print_summary(df: pd.DataFrame, horizon: str, series: str):
    from scipy import stats
    print(f"\n--- {series} series | n={len(df)} quarters ---")
    methods = [("am", "AM"), ("bc", "BC"), ("lp", "Log pool"), ("tm", "Trimmed mean")]
    for metric, label, better in [("rps", "RPS", "lower"), ("ls", "log score", "higher"),
                                   ("mass", "P(correct bin)", "higher")]:
        if f"am_{metric}" not in df.columns:
            continue
        print(f"\n  {label} ({better} is better):")
        for col_prefix, name in methods:
            col = f"{col_prefix}_{metric}"
            if col not in df.columns:
                continue
            print(f"    {name:<14} mean={df[col].mean():.4f}")
        # Wilcoxon + DM tests vs AM
        am_col = f"am_{metric}"
        for col_prefix, name in methods[1:]:
            col = f"{col_prefix}_{metric}"
            if col not in df.columns:
                continue
            diff = df[col] - df[am_col]
            alt  = "less" if better == "lower" else "greater"
            try:
                _, wpval = stats.wilcoxon(diff, alternative=alt)
            except Exception:
                wpval = float("nan")
            wins = (diff < 0).mean() * 100 if better == "lower" else (diff > 0).mean() * 100
            if better == "lower":
                _, dmpval = _dm_test(df[col], df[am_col])
            else:
                _, dmpval = _dm_test(df[am_col], df[col])
            print(f"    {name:<14} wins {wins:.1f}%  Wilcoxon p={wpval:.4f}  DM p={dmpval:.4f} vs AM")


def _plot_scores(df: pd.DataFrame, horizon: str, series: str):
    fig, axes = plt.subplots(1, 2, figsize=(12, 4))
    for ax, (am_col, bc_col, label) in zip(axes, [
        ("am_rps", "bc_rps", "RPS (lower = better)"),
        ("am_ls",  "bc_ls",  "Log Score (higher = better)"),
    ]):
        ax.plot(df["year"] + (df["qtr"] - 1) / 4, df[am_col],
                color="#d62728", label="Arithmetic Mean", alpha=0.8, lw=1.5)
        ax.plot(df["year"] + (df["qtr"] - 1) / 4, df[bc_col],
                color="#1f77b4", label="W1 Barycenter", alpha=0.8, lw=1.5)
        ax.set_xlabel("Survey quarter")
        ax.set_ylabel(label)
        ax.legend(fontsize=9)
        ax.set_title(label)
    fig.suptitle(f"SPF backtest — {horizon}-year horizon, {series} realizations",
                 fontsize=11)
    fig.tight_layout()
    fname = OUT_DIR / f"scores_{horizon}_{series}.pdf"
    fig.savefig(fname, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved {fname.name}")


def _plot_dispersion_split(df: pd.DataFrame, horizon: str, series: str):
    med_disp = df["dispersion"].median()
    df = df.copy()
    df["hi_disp"] = df["dispersion"] >= med_disp

    fig, axes = plt.subplots(1, 2, figsize=(10, 4))
    for ax, group, label in zip(axes, [True, False],
                                ["High dispersion", "Low dispersion"]):
        sub = df[df["hi_disp"] == group]
        categories = ["AM", "BC"]
        rps_means  = [sub["am_rps"].mean(), sub["bc_rps"].mean()]
        ax.bar(categories, rps_means, color=["#d62728", "#1f77b4"], alpha=0.8)
        ax.set_title(f"{label} (n={len(sub)})")
        ax.set_ylabel("Mean RPS")
    fig.suptitle(f"RPS by dispersion — {horizon}-year, {series}", fontsize=11)
    fig.tight_layout()
    fname = OUT_DIR / f"dispersion_{horizon}_{series}.pdf"
    fig.savefig(fname, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved {fname.name}")


def _plot_reliability(df: pd.DataFrame, series: str):
    """
    Bin-level reliability diagram.
    For each bin k and each method, plot mean predicted probability
    against empirical frequency of that bin being realized.
    """
    methods = [
        ("am", "Arith. mean", "#d62728"),
        ("bc", "W1 Barycenter", "#1f77b4"),
        ("lp", "Log pool", "#2ca02c"),
        ("tm", "Trimmed mean", "#ff7f0e"),
    ]
    n_total = len(df)
    fig, axes = plt.subplots(1, len(methods), figsize=(14, 4), sharey=True)
    for ax, (prefix, name, color) in zip(axes, methods):
        mean_pred = np.array([df[f"{prefix}_p{k}"].mean() for k in range(K)])
        obs_freq  = np.array([(df["bin_idx"] == k).mean() for k in range(K)])
        ax.scatter(mean_pred, obs_freq, color=color, s=50, zorder=3)
        for k in range(K):
            ax.annotate(BIN_LABELS[k], (mean_pred[k], obs_freq[k]),
                        textcoords="offset points", xytext=(4, 2), fontsize=6)
        lim = max(mean_pred.max(), obs_freq.max()) * 1.1
        ax.plot([0, lim], [0, lim], "k--", lw=0.8, label="45° (perfect)")
        ax.set_xlabel("Mean predicted probability")
        ax.set_title(name)
        ax.set_xlim(0, lim)
        ax.set_ylim(0, lim)
    axes[0].set_ylabel("Empirical frequency")
    fig.suptitle(f"Reliability diagram — SPF current-year, {series} ({n_total} quarters)",
                 fontsize=10)
    fig.tight_layout()
    fname = OUT_DIR / f"fig_bt8_reliability_{series}.pdf"
    fig.savefig(fname, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved {fname.name}")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Run both horizons
    df_cur  = run_backtest("current")
    df_next = run_backtest("next")

    print("\nDone.  All output in output/spf_backtest/")
