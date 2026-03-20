"""
Fig bt5: Fan chart showing AM and BC forecast distributions over time,
with realized Q4/Q4 core CPI overlaid.

For each quarter we reconstruct the full AM and BC PMF, then compute:
  - Expected value (mean)
  - 25th and 75th percentile (from the CDF)
  - 10th and 90th percentile

Realized values come from the existing backtest_current.csv.

Interpreter: /Users/aarondanielson/miniforge3/envs/supercluster/bin/python3.11
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from pathlib import Path
import openpyxl

ROOT    = Path(__file__).parent.parent.parent
DATA_PATH = ROOT / "data" / "spf" / "SPFmicrodata.xlsx"
OUT_DIR = ROOT / "output" / "spf_backtest"

# ── Style ─────────────────────────────────────────────────────────────────────
plt.rcParams.update({
    "font.family":      "serif",
    "font.size":        11,
    "axes.spines.top":  False,
    "axes.spines.right":False,
    "axes.linewidth":   0.8,
    "xtick.major.size": 4,
    "ytick.major.size": 4,
    "pdf.fonttype":     42,
    "ps.fonttype":      42,
})

AM_COLOR = "#c0392b"
BC_COLOR = "#2471a3"

BIN_MIDPOINTS = np.array([-0.5, 0.2, 0.75, 1.25, 1.75, 2.25, 2.75, 3.25, 3.75, 4.5])
K = 10


def quarter_float(year, qtr):
    return year + (qtr - 1) / 4


def load_quarter(ws, header, year, qtr, col_slice, tol=2.0):
    prob_cols = [c for c in header if c and str(c).startswith("PRCCPI")]
    sel_cols  = prob_cols[col_slice]
    valid = []
    for row in ws.iter_rows(values_only=True):
        if row[0] != year or row[1] != qtr:
            continue
        vals = []
        for c in sel_cols:
            v = row[header.index(c)]
            if v in (None, "#N/A", ""):
                vals.append(0.0)
            else:
                try:
                    vals.append(float(v))
                except (TypeError, ValueError):
                    vals.append(0.0)
        s = sum(vals)
        if abs(s - 100.0) <= tol and s > 0:
            arr = np.array(vals) / s
            valid.append(arr[::-1])
    if len(valid) < 5:
        return None
    return np.array(valid)


def cdf_median(pmfs):
    cdfs = np.cumsum(pmfs, axis=1)
    med  = np.median(cdfs, axis=0)
    pmf  = np.diff(np.concatenate([[0.0], med]))
    pmf  = np.maximum(pmf, 0.0)
    pmf /= pmf.sum()
    return pmf


def pmf_stats(pmf):
    """Return (mean, p10, p25, p75, p90) using bin midpoints and CDF interpolation."""
    mean = float(pmf @ BIN_MIDPOINTS)
    cdf  = np.cumsum(pmf)
    # Linear interpolation within the bin that crosses the quantile
    def quantile(q):
        # bin edges (approximate): use midpoints shifted by half-bin
        # BIN_MIDPOINTS give the center; edges are roughly at midpoints ± 0.25
        # We interpolate between adjacent midpoints
        idx = np.searchsorted(cdf, q)
        idx = min(idx, K - 1)
        if idx == 0:
            return BIN_MIDPOINTS[0]
        lo_cdf = cdf[idx - 1]
        hi_cdf = cdf[idx]
        lo_val = BIN_MIDPOINTS[idx - 1]
        hi_val = BIN_MIDPOINTS[idx]
        if hi_cdf == lo_cdf:
            return lo_val
        frac = (q - lo_cdf) / (hi_cdf - lo_cdf)
        return lo_val + frac * (hi_val - lo_val)
    return mean, quantile(0.10), quantile(0.25), quantile(0.75), quantile(0.90)


# ── Load SPF microdata ────────────────────────────────────────────────────────
print("Loading SPF microdata...")
wb = openpyxl.load_workbook(DATA_PATH, read_only=True)
ws = wb["PRCCPI"]
rows_all = list(ws.iter_rows(values_only=True))
header   = rows_all[0]
wb.close()

quarters = sorted(set(
    (int(r[0]), int(r[1])) for r in rows_all[1:] if r[0] is not None
))

col_slice = slice(0, 10)  # current-year

# ── Load realized values from existing backtest CSV ───────────────────────────
bt = pd.read_csv(OUT_DIR / "backtest_current.csv")
bt = bt[bt["series"] == "SA"].set_index(["year", "qtr"])

# ── Compute PMF stats per quarter ─────────────────────────────────────────────
records = []
for (yr, qt) in quarters:
    if yr < 2007:
        continue
    if (yr, qt) not in bt.index:
        continue

    # Reconstruct from rows_all (avoid re-opening workbook)
    prob_cols = [c for c in header if c and str(c).startswith("PRCCPI")]
    sel_cols  = prob_cols[col_slice]
    valid = []
    for row in rows_all[1:]:
        if row[0] != yr or row[1] != qt:
            continue
        vals = []
        for c in sel_cols:
            v = row[header.index(c)]
            if v in (None, "#N/A", ""):
                vals.append(0.0)
            else:
                try:
                    vals.append(float(v))
                except (TypeError, ValueError):
                    vals.append(0.0)
        s = sum(vals)
        if abs(s - 100.0) <= 2.0 and s > 0:
            arr = np.array(vals) / s
            valid.append(arr[::-1])
    if len(valid) < 5:
        continue

    pmfs = np.array(valid)
    am   = pmfs.mean(axis=0)
    bc   = cdf_median(pmfs)

    am_stats = pmf_stats(am)
    bc_stats = pmf_stats(bc)
    realized = bt.loc[(yr, qt), "realized"]

    records.append({
        "t": quarter_float(yr, qt),
        "year": yr, "qtr": qt,
        "realized": realized,
        "am_mean": am_stats[0], "am_p10": am_stats[1], "am_p25": am_stats[2],
        "am_p75": am_stats[3], "am_p90": am_stats[4],
        "bc_mean": bc_stats[0], "bc_p10": bc_stats[1], "bc_p25": bc_stats[2],
        "bc_p75": bc_stats[3], "bc_p90": bc_stats[4],
    })

df = pd.DataFrame(records).sort_values("t").reset_index(drop=True)
df.to_csv(OUT_DIR / "backtest_distributions.csv", index=False)
print(f"  {len(df)} quarters processed")


# ── Fig 5: Fan chart ──────────────────────────────────────────────────────────

fig, axes = plt.subplots(2, 1, figsize=(10, 7), sharex=True,
                         gridspec_kw={"hspace": 0.08})

for ax, prefix, color, label in [
    (axes[0], "am", AM_COLOR, "Arithmetic mean"),
    (axes[1], "bc", BC_COLOR, "W$_1$ barycenter"),
]:
    t = df["t"].values

    # 10–90% band
    ax.fill_between(t, df[f"{prefix}_p10"], df[f"{prefix}_p90"],
                    color=color, alpha=0.12, lw=0, label="10–90th pctile")
    # 25–75% band (IQR)
    ax.fill_between(t, df[f"{prefix}_p25"], df[f"{prefix}_p75"],
                    color=color, alpha=0.25, lw=0, label="25–75th pctile (IQR)")
    # Mean line
    ax.plot(t, df[f"{prefix}_mean"], color=color, lw=1.8, label="Expected value", zorder=4)
    # Realized
    ax.scatter(t, df["realized"], s=20, color="black", zorder=5,
               label="Realized Q4/Q4 core CPI" if ax is axes[0] else None)

    ax.set_ylabel("Core CPI Q4/Q4 (%)", fontsize=10)
    ax.set_title(label, fontsize=11, color=color, loc="left", pad=4)
    ax.axhline(0, color="gray", lw=0.5, ls="--", alpha=0.5)

# Shared legend
handles = [
    plt.Line2D([0], [0], color="black", lw=0, marker="o", markersize=5, label="Realized"),
    plt.Line2D([0], [0], color="gray", lw=1.8, label="Expected value"),
    mpatches.Patch(alpha=0.35, color="gray", label="IQR (25–75th)"),
    mpatches.Patch(alpha=0.20, color="gray", label="80% range (10–90th)"),
]
axes[0].legend(handles=handles, fontsize=9, frameon=False, ncol=2, loc="upper left")

axes[1].set_xlabel("Survey quarter")
fig.suptitle("SPF core CPI forecast distributions over time: AM vs.\ $W_1$ barycenter\n"
             "(shading = predictive uncertainty; dots = realized Q4/Q4 core CPI)",
             fontsize=11, y=1.01)

fig.tight_layout()
fig.savefig(OUT_DIR / "fig_bt5_fan_chart.pdf", bbox_inches="tight")
fig.savefig(OUT_DIR / "fig_bt5_fan_chart.png", dpi=200, bbox_inches="tight")
plt.close(fig)
print("Saved fig_bt5_fan_chart")
print(f"\nAll output in {OUT_DIR}")
